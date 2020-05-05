# -*- coding: utf-8 -*-
# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.4.2
#   kernelspec:
#     display_name: 'Python 3.8.2 64-bit (''data'': venv)'
#     language: python
#     name: python38264bitdatavenv86c8e9e79e034983bf9e66a89847a1a4
# ---

# %% [markdown]
# # 文件的方式处理

# %% [markdown] jupyter={"source_hidden": true} toc-hr-collapsed=true toc-nb-collapsed=true
# ## xlrd/xlwt处理excel文件
# 1. 以文件的思路处理excel数据，对于要边读边改，还要进行数据处理的需要，这种思路较为落后。

# %% execution={"iopub.execute_input": "2020-04-16T03:41:34.492Z", "iopub.status.busy": "2020-04-16T03:41:34.487Z", "iopub.status.idle": "2020-04-16T03:41:34.589Z", "shell.execute_reply": "2020-04-16T03:41:34.605Z"} jupyter={"outputs_hidden": true, "source_hidden": true}
import xlrd  # 引入模块b
import xlwt
from xlutils.copy import copy

# 打开文件，获取excel文件的workbook（工作簿）对象
workbook = xlrd.open_workbook("data/安运公交.xls")  # 文件路径
names = workbook.sheet_names()
print(names)

# %% jupyter={"outputs_hidden": true, "source_hidden": true}
worksheet = workbook.sheet_by_name("101")
print(worksheet)

nrows = worksheet.nrows  # 获取该表总行数
print(nrows)

ncols = worksheet.ncols  # 获取该表总列数
print(ncols)

# %% jupyter={"outputs_hidden": true, "source_hidden": true}
newWB = copy(workbook)
newWS = newWB.get_sheet(0)
for i in range(nrows):  # 循环打印每一行
    if (i >= 3):
        for j in range(4, 8):
            old_v = worksheet.cell_value(i, j)
            new_v = old_v.replace('.', '/')
            newWS.write(i, j, new_v)
newWB.save('data/安运公交0.xls')

# %% [markdown]
# # pandas数据的方式处理
# 不用文件的思路处理数据，用数据的思路进行处理.

# %%
import pandas as pd

pd.options.display.html.table_schema = True

df = pd.read_excel('data/安运公交 2.xls', sheet_name='101')
df2 = df['车长/宽/高（mm）'].str.split('/', expand=True)
# df3 = df.drop('车长/宽/高（mm）', axis=1).join(df2)
df2 = pd.concat([df, df2], axis=1)
df2.drop('车长/宽/高（mm）', inplace=True, axis=1)

df2.rename(columns={0: '车长mm', 1: '车宽mm', 2: '车高mm'}, inplace=True)
df2['出厂日期'] = df2['出厂日期'].str.replace('.', '/')
df2['行驶证   注册日期'] = df2['行驶证   注册日期'].str.replace('.', '/')
df2['上车日期'] = df2['上车日期'].str.replace('.', '/')
df2['行驶证         发证日期'] = df2['行驶证         发证日期'].str.replace('.', '/')
df2.to_excel('data/安运公交-拆列.xls', index=False)

# %% [markdown]
# ## 多个sheet页并拆分列（公交车数据）

# %%
import datetime
import pandas as pd

f_name = 'data/安运公交 3.xls'
print('处理文件：' + f_name)
# writer = pd.ExcelWriter('安运公交-拆列1.xlsx',engine='xlsxwriter')
writer = pd.ExcelWriter('data/安运公交-拆列@' + datetime.date.today().strftime("%Y-%m-%d") + '.xls')
xls_f = pd.ExcelFile(f_name)
for i in xls_f.sheet_names:
    # df = pd.read_excel('安运公交 3.xls',i,header=2)
    # if i.startswith('101'):
    # df_101 = df
    df = xls_f.parse(i, header=2)
    print('处理' + f_name + '的sheet页：' + i)
    df2 = df['车长/宽/高（mm）'].str.split('/', expand=True)
    df3 = df.drop('车长/宽/高（mm）', axis=1).join(df2)
    df3.rename(columns={'序 号': '公交公司', 0: '车长mm', 1: '车宽mm', 2: '车高mm'}, inplace=True)
    df3['公交公司'] = '安运公交'
    df3.to_excel(writer, sheet_name=i, index=False)
writer.save()

# %% [markdown] toc-hr-collapsed=true toc-nb-collapsed=true
# ## 同一页中多个带标题的表格处理

# %% [markdown]
#

# %% [markdown]
# ## 读取多层索引文件（勤务数据）
# 将第1列和第2列做为索引（两种方法，读取时或用set_index方法）；
# 填充excel的合并单元格（填充NaN值）；
# 按条件筛选行，并删除；
# **注意inplace=True,在原数据中直接修改生效，而无需返回copy；**

# %%
import datetime
import pandas as pd
from openpyxl import load_workbook

file_name = 'data/地铁九号线2020年4月份民警执勤表-melt.xlsx'

book = load_workbook(file_name)
writer = pd.ExcelWriter(file_name, engine='openpyxl')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

df = pd.read_excel(file_name, sheet_name=0, header=1)
df.fillna(method='pad', inplace=True)
# 删除备注行
bz = df[df['所在站'].str.contains('备注')]
df = df.drop(bz.index)
df.set_index(['所在站', '时间'], inplace=True)

# %% [markdown]
# ## Data reshape（勤务数据）

# %% [markdown]
# ### concat 不同列的数据到一列

# %% [markdown]
# 读取多层索引文件（勤务数据）,
# 将第1列和第2列做为索引（两种方法，读取时或用set_index方法）；
# 填充excel的合并单元格（填充NaN值）；
# 按条件筛选行，并删除；
# **注意inplace=True的使用；

# %%
df1 = df.iloc[:, [0, 1, 2]]
# df2 = df.iloc[:,[3,4,5]].copy()
# df2.rename(columns={'日期.1':'日期','姓名.1':'姓名','电话.1':'电话'},inplace=True)
df2 = df.iloc[:, [3, 4, 5]].rename(columns={'日期.1': '日期', '姓名.1': '姓名', '电话.1': '电话'}, inplace=False)
# df3 = df.iloc[:,[6,7,8]].copy()
# df3.rename(columns={'日期.2':'日期','姓名.2':'姓名','电话.2':'电话'},inplace=True)
df3 = df.iloc[:, [6, 7, 8]].rename(columns={'日期.2': '日期', '姓名.2': '姓名', '电话.2': '电话'}, inplace=False)
df_new = pd.concat([df1, df2, df3])

# print(df_new)
df_new.reset_index(inplace=True)
print(df_new)

df_new.to_excel(writer, sheet_name='pb_new', index=False)
writer.save()
writer.close()

# %% [markdown]
# ### melt数据（勤务数据）
# melt不可行，还是用concat。

# %% [markdown]
# ## 探索重复数据（出租车数据）

# %% [markdown]
# ### 数据探索和验证

# %% [markdown]
# 探索过程：以出租车数据为例；在出租车数据中找到脏的数据;
#     这个case中，不断的探索数据，理解数据，根据数据的条数先发现了数据存在的问题，sheet页中共有21012条数据，通过nunique对列操作，发现不重复的车牌数据有19443条，重复的车牌有1563条，浏览数据时发现有一条无关的数据“业务名称”，这样还差5条数据，在去除完全重复的数据后，对车牌series做value_counts()操作，发现还有为2的，说明有多次重复的。在merge之后找到重复的车牌数据后进行value_count()操作，发现还有重复3次的数据，6条，最终目的是找到这6条数据。
#     主要熟悉使用如下API，并注意：
#     1. 数据探索：在df的列上进行nunique()操作，在df的列上进行value_counts()操作，查看重复的情况可以连续两次value_counts()操作；value_counts()只能在series上操作；
#     2. df上进行duplicate()操作；    \n",
#     3. df上进行drop_duplicates()操作，主要在列上操作返回的是series，如果不想返回series，则在参数中指定subset；彻底熟悉keep的用法；
#     4. 如何没有列名，最好指定列名；
#     5. 看df有多少数据，用len(df)；
#     6. 初步使用merge，merge就是SQL的join，对两个datafrme的重复数据可以进行筛选；
#     7. 数据量比较大时，输出到excel里看中间的数据情况；
#     8. 筛选重复的数据，如a,b,c,d,b,c,d，需要计算出重复的有3条数据；

# %% [markdown]
# 第一个需求，探索数据，把不重复的和重复的数据量找出来，和预想的进行比较，看是不是相同；

# %%
import datetime
import pandas as pd
from openpyxl import load_workbook

file_name = 'data/2019.11.20出租企业统计表.xlsx'
df = pd.read_excel(file_name, sheet_name='车辆数统计', header=None)
print(df.head())
print(df.count())

print(df[0].nunique())
print(df[1].nunique())

print('------------------------------------')
a = df.drop_duplicates(keep='first')
print('去除完全重复的，但保留重复的第一个：', len(a))
print(a[1].value_counts().value_counts())
b = df.drop_duplicates(keep=False)
print('只要完全重复的就删除：', len(b))
a = a.append(b).drop_duplicates(keep=False)
print('完全重复的数量：', len(a))

a1 = df.drop_duplicates(subset=0, keep='first')
print('公司，去除重复的，但保留重复的第一个：', len(a1))
b1 = df.drop_duplicates(subset=0, keep=False)
print('公司，只要重复的就删除（这个值重要，只要不重复的（17880）+重复的（3132）=总量就对了）：', len(b1))
a1 = a1.append(b1).drop_duplicates(subset=0, keep=False)
print('公司，重复的数量：', len(a1))
# print('+++个体经营车辆：',df[df[0].str.contains('个体')][1].nunique())
# print('+++公司经营车辆：',df[df[0].str.contains('公司')][1].nunique())

a1 = df[1].drop_duplicates(keep='first')
print('车牌，去除重复的，但保留重复的第一个：', len(a1))
b1 = df[1].drop_duplicates(keep=False)
print('车牌，只要重复的就删除：', len(b1))
a1 = pd.Series.to_frame(a1.append(b1).drop_duplicates(keep=False))
print('车牌，重复的数量：', len(a1))
print('------------------------------------')

df.columns = ['gs', 'cp']
a1.columns = ['cp']

a2 = pd.merge(df, a1, on='cp', how='inner')
# writer = pd.ExcelWriter('data/重复车牌的数据@' + datetime.date.today().strftime("%Y-%m-%d") + '.xls')
# a2.to_excel(writer)
# writer.save()
# print('+++个体经营重复车牌：',a2[a2['gs'].str.contains('个体')]['cp'].nunique())

print('重复的车牌数据共：', len(a2))
print('重复车牌中独立的车牌有：', a2['cp'].nunique())
print(a2['cp'].value_counts())
print('重复车牌中，重复的次数（这个数据值探索重要）：')
print(a2['cp'].value_counts().value_counts())
# a2 = a2.drop_duplicates(subset='cp',keep=None)

# %% [markdown]
# ### 筛选出重复的数据
# TODO: 200417:第二个需求是，找出重复车牌，完全重复的有哪些，同一车牌被分散在不同公司的数据有哪些；
# %%
# 200506:
a2.drop_duplicates(keep=False)
# %% [markdown]
# ## pandas-profiling

# %% [markdown]
# ### profile 出租车车辆数统计profile

# %% jupyter={"outputs_hidden": true}
import datetime
import pandas as pd
import pandas_profiling

file_name = 'data/2019.11.20出租企业统计表.xlsx'
df = pd.read_excel(file_name, sheet_name='车辆数统计', header=None)
profile = df.profile_report(title='Taxi')
profile.to_file(output_file='profilings/出租企业统计表_车辆数统计报告.html')

# %% [markdown]
# poi-沈阳市-150700公交站点统计profile

# %%
import datetime
import pandas as pd
import pandas_profiling

file_name = 'data/poi-沈阳市-150700.xlsx'
df = pd.read_excel(file_name, sheet_name='150700', header=0)
gp = df.groupby(by=['lon', 'lat'])
print('按经纬度分组后统计数量（即位置不同的车站数量）:%s' % len(gp.size()))
df_loc_distinct = df[['lon', 'lat']].drop_duplicates(keep='first')
print('经纬度去重后，不同经纬度的车站有:%s个' % len(df_loc_distinct))
profile = df.profile_report(title='name')
profile.to_file(output_file='profilings/poi-沈阳市-150700公交站点报告.html')

# %% [markdown] nteract={"transient": {"deleting": false}}
# ## pyechart & pandas（客流量数据）

# %% nteract={"transient": {"deleting": false}} outputExpanded=false
import pandas as pd
import numpy as np
import time

from pyecharts.charts import Line
from pyecharts.charts import Bar
from pyecharts import options as opts
from pyecharts.globals import CurrentConfig, NotebookType
CurrentConfig.NOTEBOOK_TYPE = NotebookType.JUPYTER_LAB

from __future__ import unicode_literals
from snapshot_pyppeteer import snapshot
from pyecharts.render import make_snapshot


def convertTime(x):
    y = time.localtime(x / 1000)
    z = time.strftime("%Y-%m-%d %H:%M:%S", y)
    return z


# %% [markdown]
# 装载客流量统计表的数据

# %%
months = ['二月', '三月', '四月']
dfs = {i: pd.read_excel('data/客流量统计表.xls', sheet_name=i, header=1) for i in months}
print(dfs.keys())

# %% [markdown]
# ### 尝试pyechart示例
#
# 使用用pandas的数据代入，series类型的值可以直接用，loc返回的是dataframe，需做转换。

# %%
df = dfs['四月']
bar = (Bar().add_xaxis([2011, 2012, 2013, 2014, 2015, 2016,
                        2017]).add_yaxis("产品销量", df.loc[0:6, '地铁（人次）'].tolist()).set_global_opts(
                            title_opts=opts.TitleOpts(title="11 ~ 17年 xxx 公司 xx 产品销量图", subtitle="这里是副标题")))
bar.load_javascript()

# %%
bar.render_notebook()
# bar.render(path="charts/pyechart-snapshot.png",delay=3, pixel_ratio=3)
# make_snapshot(snapshot, bar.render(), "charts/bar.png")

# %% [markdown]
# ### 尝试eplot

# %% [markdown]
# 尝试eplot，试了一下，但是更多的参数不知道怎么用，无法进一步控制图中显示的内容。

# %%
from eplot import eplot
import numpy as np
import pandas as pd
# data环境需要安装eplot包
df1 = pd.DataFrame([
    np.random.uniform(10, 1, size=1000),
    np.random.uniform(10, 5, size=1000),
    np.random.randint(1, high=10, size=1000),
    np.random.choice(list('ABCD'), size=1000)
],
                   index=['col1', 'col2', 'col3', 'col4']).T
df1.eplot.line()

# %% [markdown]
# ### 折线图（客流总量数据）

# %% [markdown] nteract={"transient": {"deleting": false}}
# 数据探索，一是看excel的数据，二是进行数据预处理：
# 1. 删除备注；
# 2. 二月的填充NaN为0；
# 3. 三月四月的删除有空值的行；

# %% nteract={"transient": {"deleting": false}} outputExpanded=false
for month in dfs.keys():
    df = dfs[month]
    df.drop('备注', axis=1, inplace=True)
    if (month == '二月'):
        df.fillna(0, inplace=True)
        # print(df)
    else:
        df.drop(df[df.isnull().T.any()].index, inplace=True)
        df['日期'] = pd.to_datetime(df['日期'], format="%Y-%m-%d")
        # df['日期']=df['日期'].apply(convertTime)
        # df['日期']=pd.to_datetime(df['日期'])

# %% [markdown]
# pyechart画折线图

# %%
for month in dfs.keys():
    df = dfs[month]
    line = (Line().set_global_opts(
        title_opts=opts.TitleOpts(title="2020年%s地铁公交长客客流" % month),
        toolbox_opts=opts.ToolboxOpts(
            is_show=False,
            orient='vertical',
            pos_top=40,
            pos_left=5,
            feature=opts.ToolBoxFeatureOpts(
                save_as_image=opts.ToolBoxFeatureSaveAsImageOpts(type_='jpeg', pixel_ratio=2))),
        xaxis_opts=opts.AxisOpts(type_="category"),
        yaxis_opts=opts.AxisOpts(
            type_="value",
            axistick_opts=opts.AxisTickOpts(is_show=True),
            splitline_opts=opts.SplitLineOpts(is_show=True),
        ),
    ).add_xaxis(xaxis_data=['{} 日'.format(i) for i in df['日期'].dt.day]).add_yaxis(
        series_name="地铁",
        y_axis=df['地铁（人次）'],
        symbol="emptyCircle",
        is_symbol_show=True,
        color="#6e9ef1",
        label_opts=opts.LabelOpts(is_show=False),
    ).add_yaxis(
        series_name="公交",
        y_axis=df['公交（人次）'],
        symbol="emptyCircle",
        is_symbol_show=True,
        label_opts=opts.LabelOpts(is_show=False),
        markpoint_opts=opts.MarkPointOpts(data=[opts.MarkPointItem(type_="max", name="最高流量")], symbol_size=90),
    ).add_yaxis(
        series_name="长客",
        y_axis=df['长客（人次）'],
        symbol="emptyCircle",
        is_symbol_show=True,
        label_opts=opts.LabelOpts(is_show=False),
    ))
    line.render('charts/%s客流图.html' % month)
    # if(month=='四月'):
    #    line.render('charts/四月测试.png')

# %% [markdown]
# pychart输出图片共用尝试了三个工具（其实背后的原理都是模拟调用chromium）：
# 1. pyechart官方snapshot-pyppeteer 0.0.2.
# 程序中运行长时间没有结果（还需看下问题在哪）,不能在循环中运行，每次抓取没有那么快。
# 官方文档推荐的方式，更新间隔时间最短，snapshot的方式除pyppeteer（使用chromium）之外，还有selemium，phantomjs之类的自动化模拟工具，只试了pyppeteer。
# >
# from snapshot_pyppeteer import snapshot
# from pyecharts.render import make_snapshot
# make_snapshot(snapshot, line.render(), '11.png', notebook=True)
#
# 2. pyecharts-snapshot 0.2.0，命令行可以运行输出静态图，一年没更新了。
# >
# from __future__ import unicode_literals
# line.render('charts/四月测试.png')
# 运行后的图片无法打开.
# 命令行：
# snapshot 四月客流图.html
#
# 前两者是python包，还有一种是npm包，是可以将echart输出的html进行输出，也支持pyechart输出的
#
# 3. echarts-scrappeteer
# 介绍中可以输出gif动态图，但还没有搞明参数怎么使用。两年多了没更新了。
# 在命令行中调用时，会蹦出chromium。
# >
# scrappeteer
#

# %% [markdown] toc-hr-collapsed=true toc-nb-collapsed=true
# # Dataframe写入MySQL的表

# %% jupyter={"outputs_hidden": true}
